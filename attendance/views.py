from datetime import datetime
from io import BytesIO
from typing import Optional

from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import render, redirect
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.utils import timezone

from .models import Employee, Attendance, Shift

# ---------- 小ユーティリティ ----------
def _employee_queryset():
    qs = Employee.objects.all().order_by("code")
    try:
        active = Employee.objects.filter(is_active=True).order_by("code")
        if active.exists():
            qs = active
    except Exception:
        pass
    return qs

def _today_attendance(employee: Employee) -> Attendance:
    # モデルは work_date を使う前提
    today = timezone.localdate()
    att, _ = Attendance.objects.get_or_create(employee=employee, work_date=today)
    return att

# ---------- サービス層のロード（あれば優先） ----------
try:
    from codereview.attendance.services import PunchService
except Exception:
    try:
        from .services import PunchService
    except Exception:
        class PunchService:  # フォールバック（最小）
            def __init__(self, employee: Employee, note: str = "") -> None:
                self.employee = employee
                self.note = note
            def punch_in(self):
                att = _today_attendance(self.employee)
                if att.time_in is None:
                    att.time_in = timezone.localtime().time()
                    att.save()
            def punch_out(self):
                att = _today_attendance(self.employee)
                if att.time_in is None:
                    raise ValueError("先に出勤を打刻してください。")
                if att.time_out is None:
                    att.time_out = timezone.localtime().time()
                    att.save()

def _build_punch_service(employee: Employee, note: str):
    try:
        return PunchService(employee, note=note)
    except TypeError:
        return PunchService(employee)

def _call_punch(svc, action: str):
    # punch_* / clock_* どちらでも対応
    if action == "in":
        if hasattr(svc, "punch_in"):
            svc.punch_in()
        elif hasattr(svc, "clock_in"):
            svc.clock_in(timezone.localtime().time())
        else:
            att = _today_attendance(svc.employee)
            if att.time_in is None:
                att.time_in = timezone.localtime().time()
                att.save()
    elif action == "out":
        if hasattr(svc, "punch_out"):
            svc.punch_out()
        elif hasattr(svc, "clock_out"):
            svc.clock_out(timezone.localtime().time())
        else:
            att = _today_attendance(svc.employee)
            if att.time_in is None:
                raise ValueError("先に出勤を打刻してください。")
            if att.time_out is None:
                att.time_out = timezone.localtime().time()
                att.save()
    else:
        raise ValueError("不明な操作です。")

# ---------- ここからビュー ----------
@require_http_methods(["GET", "POST"])
def punch(request):
    # フォーム依存をこの関数の中に限定（インポート失敗で全体が落ちないように）
    from .forms import PunchForm

    if request.method == "POST":
        form = PunchForm(request.POST)
        form.fields["employee"].queryset = _employee_queryset()
        if not form.is_valid():
            messages.error(request, "入力内容を確認してください。")
            return render(request, "attendance/punch.html", {"form": form})

        employee: Employee = form.cleaned_data["employee"]
        note = (request.POST.get("note") or "").strip()
        action = (request.POST.get("action") or "").strip().lower()

        try:
            svc = _build_punch_service(employee, note)
            _call_punch(svc, action)
            messages.success(
                request,
                f"{getattr(employee, 'name', str(employee))}："
                + ("出勤を記録しました。" if action == "in" else "退勤を記録しました。")
            )
        except Exception as e:
            messages.error(request, f"打刻に失敗しました: {e}")
        return redirect("punch_page")

    # GET
    form = PunchForm()
    form.fields["employee"].queryset = _employee_queryset()
    return render(request, "attendance/punch.html", {"form": form})

def healthcheck(request):
    return HttpResponse("OK")

# ---- 以下は base.html の URL 逆引き対策（プレースホルダ中心）----
def attendance_list(request):
    # 最小表示（テンプレがあればそちらを使う）
    try:
        qs = Attendance.objects.select_related("employee").order_by("-work_date", "employee__code")[:200]
        return render(request, "attendance/attendance_list.html", {"attendances": qs})
    except Exception:
        return HttpResponse("attendance_list placeholder")

def attendance_search(request):
    try:
        q = (request.GET.get("q") or "").strip()
        qs = Attendance.objects.select_related("employee")
        if q:
            from django.db.models import Q
            qs = qs.filter(Q(employee__code__icontains=q) | Q(employee__name__icontains=q) | Q(note__icontains=q))
        qs = qs.order_by("-work_date", "employee__code")[:500]
        return render(request, "attendance/attendance_search_results.html", {"attendances": qs, "query": q})
    except Exception:
        return HttpResponse("attendance_search placeholder")

def shift_list(request):
    try:
        ym = (request.GET.get("month") or timezone.localdate().strftime("%Y-%m"))
        emp_query = (request.GET.get("emp") or "").strip()
        qs = Shift.objects.select_related("employee")
        if emp_query:
            from django.db.models import Q
            qs = qs.filter(Q(employee__code__icontains=emp_query) | Q(employee__name__icontains=emp_query))
        dt = datetime.strptime(ym + "-01", "%Y-%m-%d").date()
        from calendar import monthrange
        last_day = monthrange(dt.year, dt.month)[1]
        qs = qs.filter(date__range=[dt.replace(day=1), dt.replace(day=last_day)]).order_by("date", "employee__code")
        return render(request, "attendance/shift_list.html",
                      {"shifts": qs, "ym": ym.replace("-", ""), "month_value": ym, "emp_query": emp_query})
    except Exception:
        return HttpResponse("shift_list placeholder")

def shift_create(request):
    try:
        from .forms import ShiftForm
        form = ShiftForm(request.POST or None)
        if request.method == "POST" and form.is_valid():
            form.save()
            messages.success(request, "シフトを登録しました。")
            return redirect("shift_list")
        return render(request, "attendance/shift_form.html", {"form": form})
    except Exception:
        return HttpResponse("shift_create placeholder")

def employee_register(request):
    try:
        from .forms import EmployeeForm
        form = EmployeeForm(request.POST or None)
        if request.method == "POST" and "add-submit" in request.POST:
            if form.is_valid():
                form.save()
                messages.success(request, "従業員を登録しました。")
                return redirect("employee_register")
            messages.error(request, "入力内容を確認してください。")
        return render(request, "attendance/employee_register.html", {"form": form})
    except Exception:
        return HttpResponse("employee_register placeholder")

def employee_delete(request):
    return HttpResponse("employee_delete placeholder")

def import_hub(request):
    try:
        from .forms import EmployeeImportForm, ShiftImportForm
        ctx = {"emp_form": EmployeeImportForm(), "shift_form": ShiftImportForm()}
        return render(request, "attendance/import_hub.html", ctx)
    except Exception:
        return HttpResponse("import_hub placeholder")

def employee_import(request):
    return HttpResponse("employee_import placeholder")

def shift_import(request):
    return HttpResponse("shift_import placeholder")

def download_hub(request):
    try:
        return render(request, "attendance/download_hub.html")
    except Exception:
        return HttpResponse("download_hub placeholder")

def download_employees(request):
    try:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws.append(["code","name","hourly_rate","is_active"])
        for e in Employee.objects.all().order_by("code"):
            ws.append([e.code, e.name, float(e.hourly_rate or 0), "1" if e.is_active else "0"])
        return _xlsx_response(wb, "employees.xlsx")
    except Exception:
        return HttpResponse("employee_export placeholder")

def download_shifts(request):
    try:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws.append(["code","date","start","end","note"])
        for s in Shift.objects.select_related("employee").all().order_by("date","employee__code"):
            ws.append([s.employee.code, s.date.isoformat(),
                       (s.start_time.strftime("%H:%M") if s.start_time else ""),
                       (s.end_time.strftime("%H:%M") if s.end_time else ""),
                       s.note or ""])
        return _xlsx_response(wb, "shifts.xlsx")
    except Exception:
        return HttpResponse("shift_export placeholder")

def download_attendance(request):
    try:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws.append(["date","code","name","in","out","note"])
        for a in Attendance.objects.select_related("employee").all().order_by("work_date","employee__code"):
            ws.append([a.work_date.isoformat(), a.employee.code, a.employee.name,
                       a.time_in.isoformat() if a.time_in else "",
                       a.time_out.isoformat() if a.time_out else "", a.note or ""])
        return _xlsx_response(wb, "attendance.xlsx")
    except Exception:
        return HttpResponse("attendance_export placeholder")

def _xlsx_response(wb, filename: str) -> HttpResponse:
    from openpyxl import Workbook
    if wb is None:
        from openpyxl import Workbook as _WB
        wb = _WB()
    fp = BytesIO()
    wb.save(fp)
    fp.seek(0)
    resp = HttpResponse(fp.read(),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
