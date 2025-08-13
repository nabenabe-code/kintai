from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, time
from io import BytesIO
from typing import Iterable, List, Tuple

from django.db import transaction
from django.utils import timezone

from openpyxl import Workbook, load_workbook

from .models import Attendance, Employee, Shift
@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    errors: List[str] = None

    def __post_init__(self):
        if self.errors is None:
            self.errors = []


class PunchService: #打刻
    def __init__(self, *, employee: Employee, note: str | None = None):
        self.employee = employee
        self.note = (note or "").strip()

    def _now(self):
        return timezone.localtime()

    def _today(self):
        return timezone.localdate()

    def _get_or_create_today(self) -> Attendance:
        att, _ = Attendance.objects.get_or_create(
            employee=self.employee,
            work_date=self._today(),
        )
        if self.note:
            att.note = (f"{att.note} / {self.note}".strip(" /")) if att.note else self.note
        return att

    @transaction.atomic
    def punch_in(self) -> Attendance:
        att = self._get_or_create_today()
        if att.time_in:
            raise ValueError("すでに本日の出勤が登録されています。")
        att.time_in = self._now()
        att.save(update_fields=["time_in", "note"])
        return att

    @transaction.atomic
    def punch_out(self) -> Attendance:
        att = self._get_or_create_today()
        if not att.time_in:
            raise ValueError("出勤が先に必要です。")
        if att.time_out:
            raise ValueError("すでに本日の退勤が登録されています。")
        # 退勤は出勤以降であることを簡易チェック
        now = self._now()
        if now < att.time_in:
            raise ValueError("退勤が出勤より前になっています。端末時刻を確認してください。")
        att.time_out = now
        att.save(update_fields=["time_out", "note"])
        return att
class EmployeeExcelImporter: #従業員インポート
    REQUIRED = ("code", "hourly_rate")

    def __init__(self, file_obj):
        self.file_obj = file_obj

    def _normalize_headers(self, headers: Iterable[str]) -> List[str]:
        return [str(h or "").strip().lower() for h in headers]

    def run(self) -> ImportResult:
        wb = load_workbook(self.file_obj, data_only=True)
        ws = wb.active
        result = ImportResult()

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result.errors.append("シートが空です。")
            return result

        headers = self._normalize_headers(rows[0])
        for r in self.REQUIRED:
            if r not in headers:
                result.errors.append(f"必須列がありません: {r}")
                return result

        idx_code = headers.index("code")
        idx_rate = headers.index("hourly_rate")
        idx_name = headers.index("name") if "name" in headers else None

        for i, row in enumerate(rows[1:], start=2):
            code = str(row[idx_code]).strip() if row[idx_code] is not None else ""
            rate_val = row[idx_rate]
            name = (str(row[idx_name]).strip() if (idx_name is not None and row[idx_name] is not None) else "")

            if not code:
                result.errors.append(f"{i}行目: code が空です")
                continue

            try:
                hourly_rate = float(rate_val)
            except Exception:
                result.errors.append(f"{i}行目: hourly_rate を数値にできません")
                continue

            obj, created = Employee.objects.update_or_create(
                code=code,
                defaults={"name": name, "hourly_rate": hourly_rate, "is_active": True},
            )
            if created:
                result.created += 1
            else:
                result.updated += 1

        return result


class ShiftExcelImporter: #シフトインポート

    REQUIRED = ("code", "date", "start", "end")

    def __init__(self, file_obj):
        self.file_obj = file_obj

    def _normalize_headers(self, headers: Iterable[str]) -> List[str]:
        return [str(h or "").strip().lower() for h in headers]

    def _parse_time(self, s) -> time:
        if isinstance(s, time):
            return s
        if isinstance(s, datetime):
            return s.time().replace(second=0, microsecond=0)
        s = str(s)
        hh, mm = s.split(":")
        return time(hour=int(hh), minute=int(mm))

    def run(self) -> ImportResult:
        wb = load_workbook(self.file_obj, data_only=True)
        ws = wb.active
        result = ImportResult()

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result.errors.append("シートが空です。")
            return result

        headers = self._normalize_headers(rows[0])
        for r in self.REQUIRED:
            if r not in headers:
                result.errors.append(f"必須列がありません: {r}")
                return result

        idx_code = headers.index("code")
        idx_date = headers.index("date")
        idx_start = headers.index("start")
        idx_end = headers.index("end")
        idx_note = headers.index("note") if "note" in headers else None

        for i, row in enumerate(rows[1:], start=2):
            code = str(row[idx_code]).strip() if row[idx_code] is not None else ""
            raw_date = row[idx_date]
            raw_start = row[idx_start]
            raw_end = row[idx_end]
            note = (str(row[idx_note]).strip() if (idx_note is not None and row[idx_note] is not None) else "")

            if not code:
                result.errors.append(f"{i}行目: code が空です")
                continue

            # 従業員の存在確認
            try:
                emp = Employee.objects.get(code=code, is_active=True)
            except Employee.DoesNotExist:
                result.errors.append(f"{i}行目: code={code} の従業員が見つかりません（在籍OFFの可能性含む）")
                continue

            # 日付/時刻の変換
            try:
                if isinstance(raw_date, datetime):
                    work_date = raw_date.date()
                else:
                    # 文字列 "YYYY-MM-DD" を想定
                    work_date = datetime.strptime(str(raw_date), "%Y-%m-%d").date()

                start_t = self._parse_time(raw_start)
                end_t = self._parse_time(raw_end)
            except Exception:
                result.errors.append(f"{i}行目: 日付/時刻の形式が不正です")
                continue

            obj, created = Shift.objects.update_or_create(
                employee=emp,
                work_date=work_date,
                defaults={"start_time": start_t, "end_time": end_t, "note": note},
            )
            if created:
                result.created += 1
            else:
                result.updated += 1

        return result

class ExcelExporter: # Excelエクスポート

    @staticmethod
    def _autosize(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    v = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(v))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(10, min(40, max_len + 2))

    @staticmethod
    def _to_bytes(wb: Workbook) -> bytes:
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @classmethod
    def employees(cls) -> Tuple[str, bytes]:
        wb = Workbook()
        ws = wb.active
        ws.title = "employees"
        ws.append(["社員番号", "氏名", "時給"])
        for e in Employee.objects.all().order_by("code"):
            ws.append([e.code, e.name, e.hourly_rate])
        cls._autosize(ws)
        return ("employees.xlsx", cls._to_bytes(wb))

    @classmethod
    def shifts(cls) -> Tuple[str, bytes]:
        wb = Workbook()
        ws = wb.active
        ws.title = "shifts"
        ws.append(["日付", "社員番号", "氏名", "開始", "終了", "備考"])
        qs = Shift.objects.select_related("employee").order_by("work_date", "employee__code")
        for s in qs:
            ws.append([
                s.work_date.isoformat(),
                s.employee.code,
                s.employee.name,
                s.start_time.strftime("%H:%M") if s.start_time else "",
                s.end_time.strftime("%H:%M") if s.end_time else "",
                s.note or "",
            ])
        cls._autosize(ws)
        return ("shifts.xlsx", cls._to_bytes(wb))

    @classmethod
    def attendance(cls) -> Tuple[str, bytes]:
        wb = Workbook()
        ws = wb.active
        ws.title = "attendance"
        ws.append(["勤務日", "社員番号", "氏名", "出勤", "退勤", "勤務時間[h]", "時給", "支給額", "備考"])
        qs = Attendance.objects.select_related("employee").order_by("work_date", "employee__code")
        for a in qs:
            hours = (a.work_hours or 0.0)
            rate = a.employee.hourly_rate or 0.0
            pay = round(hours * rate, 2)
            ws.append([
                a.work_date.isoformat(),
                a.employee.code,
                a.employee.name,
                a.time_in.strftime("%H:%M") if a.time_in else "",
                a.time_out.strftime("%H:%M") if a.time_out else "",
                hours,
                rate,
                pay,
                a.note or "",
            ])
        cls._autosize(ws)
        return ("attendance.xlsx", cls._to_bytes(wb))
